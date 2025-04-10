import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from werkzeug.security import generate_password_hash

from models import db, Student, User

def create_student_template():
    """Tạo file Excel mẫu cho nhập học sinh"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "StudentImportTemplate"
    
    # Define header style
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Define headers
    headers = [
        "first_name", "last_name", "email", "phone", "address", 
        "date_of_birth", "boarding_status", "meal_fee_per_day"
    ]
    
    # Add headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add example data
    example_data = [
        "Minh", "Nguyen", "minh.nguyen@example.com", "0901234567", "123 Đường Nguyễn Huệ, TP.HCM",
        "2010-05-15", "không", "50000"
    ]
    for col_idx, data in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=data)
        cell.alignment = Alignment(horizontal='left')
        cell.border = thin_border
    
    # Adjust column width
    for col_idx, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = max(len(header) * 1.5, 15)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add instruction sheet
    ws_instruction = wb.create_sheet(title="Instructions")
    ws_instruction.cell(row=1, column=1, value="Hướng dẫn nhập dữ liệu học sinh").font = Font(bold=True, size=14)
    
    instructions = [
        "1. Các trường bắt buộc: first_name, last_name, email, phone",
        "2. Định dạng ngày sinh: YYYY-MM-DD (ví dụ: 2010-05-15)",
        "3. Trạng thái bán trú: nhập một trong các giá trị 'không', 'bán trú', 'nội trú'",
        "4. Phí ăn mỗi ngày: nhập số tiền (chỉ nhập số, không nhập dấu phẩy hoặc đơn vị tiền tệ)",
        "5. Email phải là duy nhất trong hệ thống",
        "6. Xóa dòng dữ liệu mẫu trước khi nhập dữ liệu thực tế"
    ]
    
    for idx, instruction in enumerate(instructions, 3):
        ws_instruction.cell(row=idx, column=1, value=instruction)
        
    for col in range(1, 10):
        ws_instruction.column_dimensions[get_column_letter(col)].width = 20
    
    return wb

def create_teacher_template():
    """Tạo file Excel mẫu cho nhập giáo viên"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TeacherImportTemplate"
    
    # Define header style
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Define headers
    headers = [
        "username", "email", "password", "first_name", "last_name", 
        "role", "teaching_rate", "late_fee", "absence_fee"
    ]
    
    # Add headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add example data
    example_data = [
        "teacher1", "teacher1@example.com", "password123", "Van", "Tran",
        "teacher", "150000", "20000", "50000"
    ]
    for col_idx, data in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=data)
        cell.alignment = Alignment(horizontal='left')
        cell.border = thin_border
    
    # Adjust column width
    for col_idx, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = max(len(header) * 1.5, 15)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add instruction sheet
    ws_instruction = wb.create_sheet(title="Instructions")
    ws_instruction.cell(row=1, column=1, value="Hướng dẫn nhập dữ liệu giáo viên").font = Font(bold=True, size=14)
    
    instructions = [
        "1. Các trường bắt buộc: username, email, password, first_name, last_name",
        "2. Mật khẩu (password) phải có ít nhất 6 ký tự",
        "3. Vai trò (role): có thể là 'teacher', 'admin', hoặc 'sadmin'",
        "4. Tên đăng nhập (username) và email phải là duy nhất trong hệ thống",
        "5. Mức lương dạy (teaching_rate): nhập số tiền (chỉ nhập số, không nhập dấu phẩy hoặc đơn vị tiền tệ)",
        "6. Phí trừ khi đi muộn (late_fee): nhập số tiền",
        "7. Phí trừ khi vắng mặt (absence_fee): nhập số tiền",
        "8. Nếu không nhập các trường tùy chọn, hệ thống sẽ dùng giá trị mặc định",
        "9. Xóa dòng dữ liệu mẫu trước khi nhập dữ liệu thực tế"
    ]
    
    for idx, instruction in enumerate(instructions, 3):
        ws_instruction.cell(row=idx, column=1, value=instruction)
        
    for col in range(1, 10):
        ws_instruction.column_dimensions[get_column_letter(col)].width = 20
    
    return wb

def process_student_import(excel_file):
    """Xử lý file Excel để nhập danh sách học sinh"""
    try:
        df = pd.read_excel(excel_file, engine='openpyxl')
        
        # Kiểm tra các cột bắt buộc
        required_columns = ["first_name", "last_name", "email", "phone"]
        for column in required_columns:
            if column not in df.columns:
                return 0, 1, [f"Thiếu cột bắt buộc: {column}"]
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Xử lý từng dòng dữ liệu
        for index, row in df.iterrows():
            try:
                # Kiểm tra dữ liệu bắt buộc
                if pd.isna(row['first_name']) or pd.isna(row['last_name']) or pd.isna(row['email']) or pd.isna(row['phone']):
                    error_count += 1
                    errors.append(f"Dòng {index+2}: Thiếu thông tin bắt buộc")
                    continue
                
                # Kiểm tra email đã tồn tại chưa
                existing_student = Student.query.filter_by(email=row['email']).first()
                if existing_student:
                    error_count += 1
                    errors.append(f"Dòng {index+2}: Email {row['email']} đã tồn tại trong hệ thống")
                    continue
                
                # Xử lý ngày sinh
                date_of_birth = None
                if 'date_of_birth' in row and not pd.isna(row['date_of_birth']):
                    try:
                        if isinstance(row['date_of_birth'], str):
                            date_of_birth = datetime.strptime(row['date_of_birth'], '%Y-%m-%d').date()
                        else:
                            date_of_birth = row['date_of_birth'].date() if hasattr(row['date_of_birth'], 'date') else row['date_of_birth']
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Dòng {index+2}: Lỗi định dạng ngày sinh - {str(e)}")
                        continue
                
                # Xử lý trạng thái bán trú
                boarding_status = "không"
                if 'boarding_status' in row and not pd.isna(row['boarding_status']):
                    if row['boarding_status'] in ["không", "bán trú", "nội trú"]:
                        boarding_status = row['boarding_status']
                    else:
                        error_count += 1
                        errors.append(f"Dòng {index+2}: Trạng thái bán trú không hợp lệ (phải là 'không', 'bán trú', hoặc 'nội trú')")
                        continue
                
                # Xử lý phí ăn mỗi ngày
                meal_fee_per_day = 0.0
                if 'meal_fee_per_day' in row and not pd.isna(row['meal_fee_per_day']):
                    try:
                        meal_fee_per_day = float(row['meal_fee_per_day'])
                    except:
                        error_count += 1
                        errors.append(f"Dòng {index+2}: Phí ăn mỗi ngày không hợp lệ (phải là số)")
                        continue
                
                # Tạo học sinh mới
                student = Student(
                    first_name=row['first_name'],
                    last_name=row['last_name'],
                    email=row['email'],
                    phone=row['phone'],
                    address=row['address'] if 'address' in row and not pd.isna(row['address']) else None,
                    date_of_birth=date_of_birth,
                    boarding_status=boarding_status,
                    meal_fee_per_day=meal_fee_per_day,
                    active=True
                )
                
                db.session.add(student)
                db.session.commit()
                success_count += 1
                
            except Exception as e:
                db.session.rollback()
                error_count += 1
                errors.append(f"Dòng {index+2}: Lỗi - {str(e)}")
        
        return success_count, error_count, errors
    
    except Exception as e:
        return 0, 1, [f"Lỗi khi đọc file Excel: {str(e)}"]

def process_teacher_import(excel_file):
    """Xử lý file Excel để nhập danh sách giáo viên"""
    try:
        df = pd.read_excel(excel_file, engine='openpyxl')
        
        # Kiểm tra các cột bắt buộc
        required_columns = ["username", "email", "password", "first_name", "last_name"]
        for column in required_columns:
            if column not in df.columns:
                return 0, 1, [f"Thiếu cột bắt buộc: {column}"]
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Xử lý từng dòng dữ liệu
        for index, row in df.iterrows():
            try:
                # Kiểm tra dữ liệu bắt buộc
                if pd.isna(row['username']) or pd.isna(row['email']) or pd.isna(row['password']) or pd.isna(row['first_name']) or pd.isna(row['last_name']):
                    error_count += 1
                    errors.append(f"Dòng {index+2}: Thiếu thông tin bắt buộc")
                    continue
                
                # Kiểm tra tên đăng nhập đã tồn tại chưa
                existing_user_username = User.query.filter_by(username=row['username']).first()
                if existing_user_username:
                    error_count += 1
                    errors.append(f"Dòng {index+2}: Tên đăng nhập {row['username']} đã tồn tại trong hệ thống")
                    continue
                
                # Kiểm tra email đã tồn tại chưa
                existing_user_email = User.query.filter_by(email=row['email']).first()
                if existing_user_email:
                    error_count += 1
                    errors.append(f"Dòng {index+2}: Email {row['email']} đã tồn tại trong hệ thống")
                    continue
                
                # Kiểm tra độ dài mật khẩu
                if len(str(row['password'])) < 6:
                    error_count += 1
                    errors.append(f"Dòng {index+2}: Mật khẩu phải có ít nhất 6 ký tự")
                    continue
                
                # Xử lý vai trò
                role = "teacher"
                if 'role' in row and not pd.isna(row['role']):
                    if row['role'] in ["teacher", "admin", "sadmin"]:
                        role = row['role']
                    else:
                        error_count += 1
                        errors.append(f"Dòng {index+2}: Vai trò không hợp lệ (phải là 'teacher', 'admin', hoặc 'sadmin')")
                        continue
                
                # Xử lý các trường số
                teaching_rate = 0.0
                late_fee = 0.0
                absence_fee = 0.0
                
                if 'teaching_rate' in row and not pd.isna(row['teaching_rate']):
                    try:
                        teaching_rate = float(row['teaching_rate'])
                    except:
                        error_count += 1
                        errors.append(f"Dòng {index+2}: Mức lương dạy không hợp lệ (phải là số)")
                        continue
                
                if 'late_fee' in row and not pd.isna(row['late_fee']):
                    try:
                        late_fee = float(row['late_fee'])
                    except:
                        error_count += 1
                        errors.append(f"Dòng {index+2}: Phí trừ khi đi muộn không hợp lệ (phải là số)")
                        continue
                
                if 'absence_fee' in row and not pd.isna(row['absence_fee']):
                    try:
                        absence_fee = float(row['absence_fee'])
                    except:
                        error_count += 1
                        errors.append(f"Dòng {index+2}: Phí trừ khi vắng mặt không hợp lệ (phải là số)")
                        continue
                
                # Tạo người dùng mới
                user = User(
                    username=row['username'],
                    email=row['email'],
                    password_hash=generate_password_hash(str(row['password'])),
                    first_name=row['first_name'],
                    last_name=row['last_name'],
                    role=role,
                    teaching_rate=teaching_rate,
                    late_fee=late_fee,
                    absence_fee=absence_fee
                )
                
                db.session.add(user)
                db.session.commit()
                success_count += 1
                
            except Exception as e:
                db.session.rollback()
                error_count += 1
                errors.append(f"Dòng {index+2}: Lỗi - {str(e)}")
        
        return success_count, error_count, errors
    
    except Exception as e:
        return 0, 1, [f"Lỗi khi đọc file Excel: {str(e)}"]